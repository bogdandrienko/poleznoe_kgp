import datetime
import openpyxl
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from django_app import models


def index(request):
    try:
        message = ""
        if request.method == "POST":
            try:
                author = request.POST.get("author", "")
                subdivision = request.POST.get("subdivision", "")
                position = request.POST.get("position", "")
                phone = request.POST.get("phone", "")
                email = request.POST.get("email", "")
                title = request.POST.get("title", "")
                description = request.POST.get("description", "")
                place = request.POST.get("place", "")
                effect = request.POST.get("effect", "")
                need = request.POST.get("need", "")
                is_feedback = True if request.POST.get("is_feedback", "") == "да" else False
                link = request.POST.get("link", "")

                models.IdeaModel.objects.create(
                    author=author,
                    subdivision=subdivision,
                    position=position,
                    phone=phone,
                    email=email,
                    title=title,
                    description=description,
                    place=place,
                    effect=effect,
                    need=need,
                    is_feedback=is_feedback,
                    link=link,
                )
            except Exception as error:
                message = str(error)
            else:
                message = "Успешно отправлено"
        return render(request, "django_app/index.html", context={"message": message})
    except Exception as error:
        return HttpResponse(str(error))


def export(request):
    try:
        # for i in range(200, 1000):
        #     models.IdeaModel.objects.create(
        #         author=str(i) * 10,
        #         subdivision=str(i) * 10,
        #         position=str(i) * 10,
        #         phone="1111111111111111",
        #         email="email@mail.ru",
        #         title=str(i) * 100,
        #         description=str(i) * 200,
        #         place=str(i) * 200,
        #         effect=str(i) * 200,
        #         need=str(i) * 100,
        #         is_feedback=True,
        #         link=str(i) * 200,
        #     )

        ideas = models.IdeaModel.objects.all()

        workbook: Workbook = openpyxl.Workbook()
        worksheet: Worksheet = workbook.active
        titles = [
            "ФИО(полностью)",
            "Подразделение",
            "Профессия, должность",
            "Номер телефона",
            "Электронная почта",
            "Предложение(наименование)",
            "Описание",
            "Краткое описание объекта",
            "Ожидаемый эффект",
            "Необходимые ТМЦ",
            "Нужно ли связаться с работником",
            "Ссылка",
            "Дата и время создания",
        ]
        for column, title in enumerate(titles, 1):
            worksheet.cell(row=1, column=column, value=str(title))
        for row, idea in enumerate(ideas, 2):
            worksheet.cell(row=row, column=1, value=str(idea.author))
            worksheet.cell(row=row, column=2, value=str(idea.subdivision))
            worksheet.cell(row=row, column=3, value=str(idea.position))
            worksheet.cell(row=row, column=4, value=str(idea.phone))
            worksheet.cell(row=row, column=5, value=str(idea.email))
            worksheet.cell(row=row, column=6, value=str(idea.title))
            worksheet.cell(row=row, column=7, value=str(idea.description))
            worksheet.cell(row=row, column=8, value=str(idea.place))
            worksheet.cell(row=row, column=9, value=str(idea.effect))
            worksheet.cell(row=row, column=10, value=str(idea.need))
            worksheet.cell(row=row, column=11, value="нужно связаться" if idea.is_feedback else "не нужно связываться")
            worksheet.cell(row=row, column=12, value=str(idea.link))
            worksheet.cell(row=row, column=13, value=str(idea.created.strftime('%d %m %Y %H %M %S')))
        # filename = f"media/ideas/export/{datetime.datetime.now().strftime('%m_%d_%Y__%H_%M_%S')}.xlsx"
        filename = f"ideas/export/{datetime.datetime.now().strftime('%m_%d_%Y__%H_%M_%S')}.xlsx"
        workbook.save(f"/home/poleznoe/django_settings/media/{filename}")
        return render(request, "django_app/export.html", context={"message": "", "filename": filename})
    except Exception as error:
        return HttpResponse(str(error))
